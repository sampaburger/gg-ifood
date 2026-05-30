import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Resumo iFood GGK", page_icon="🍔", layout="wide")

# =========================
# Utilidades
# =========================
def read_ifood_xlsx(uploaded_file, sheet_name=None):
    """Lê arquivos XLSX exportados pelo iFood, mesmo quando a dimensão interna vem como A1."""
    data = uploaded_file.read()
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]

    try:
        ws.reset_dimensions()
    except Exception:
        pass

    rows = list(ws.iter_rows(values_only=True))
    rows = [r for r in rows if any(v is not None for v in r)]
    if not rows:
        return pd.DataFrame()

    header = [str(x).strip() if x is not None else f"col_{i}" for i, x in enumerate(rows[0])]
    return pd.DataFrame(rows[1:], columns=header)

def money(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"

def pct(v):
    try:
        return f"{float(v):.1f}%".replace(".", ",")
    except Exception:
        return "0,0%"

def num(v):
    try:
        return f"{int(round(float(v), 0)):,}".replace(",", ".")
    except Exception:
        return "0"

def to_number(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)

def sum_col(df, col, mask=None):
    if df is None or df.empty or col not in df.columns:
        return 0.0
    s = to_number(df[col])
    if mask is not None:
        s = s[mask]
    return float(s.sum())

def sum_financeiro(fin, termos, absolute=True):
    if fin.empty or "descricao_lancamento" not in fin.columns or "valor" not in fin.columns:
        return 0.0
    desc = fin["descricao_lancamento"].fillna("").astype(str).str.lower()
    mask = pd.Series(False, index=fin.index)
    for termo in termos:
        mask = mask | desc.str.contains(termo.lower(), regex=False)
    valor = to_number(fin.loc[mask, "valor"]).sum()
    return float(abs(valor) if absolute else valor)

def montar_desempenho(desempenho):
    """Extrai os dados da aba Desempenho > Vendas do iFood."""
    if desempenho is None or desempenho.empty:
        return {
            "Faturamento comercial iFood": 0.0,
            "Pedidos desempenho": 0,
            "Taxa de entrega desempenho": 0.0,
            "Ticket médio desempenho": 0.0,
            "Novos clientes": 0,
        }

    vendas_col = "Valor total de vendas"
    pedidos_col = "Total de vendas (pedidos)"
    taxa_entrega_col = "Taxa de entrega"
    novos_col = "Novos clientes"

    faturamento_comercial = sum_col(desempenho, vendas_col)
    pedidos_desempenho = sum_col(desempenho, pedidos_col)
    taxa_entrega_desempenho = sum_col(desempenho, taxa_entrega_col)
    novos_clientes = sum_col(desempenho, novos_col)
    ticket_medio_desempenho = faturamento_comercial / pedidos_desempenho if pedidos_desempenho else 0

    return {
        "Faturamento comercial iFood": faturamento_comercial,
        "Pedidos desempenho": int(round(pedidos_desempenho, 0)),
        "Taxa de entrega desempenho": taxa_entrega_desempenho,
        "Ticket médio desempenho": ticket_medio_desempenho,
        "Novos clientes": int(round(novos_clientes, 0)),
    }

def montar_resumo(pedidos, financeiro, desempenho=None):
    status_col = "STATUS FINAL DO PEDIDO"
    if status_col in pedidos.columns:
        status = pedidos[status_col].fillna("").astype(str).str.upper().str.strip()
        mask_concluido = status.eq("CONCLUIDO")
        pedidos_concluidos = int(mask_concluido.sum())
        pedidos_cancelados = int(status.eq("CANCELADO").sum())
        cancelamentos_parciais = int(status.eq("CANCELAMENTO PARCIAL").sum())
    else:
        mask_concluido = pd.Series(True, index=pedidos.index)
        pedidos_concluidos = len(pedidos)
        pedidos_cancelados = 0
        cancelamentos_parciais = 0

    pedidos_totais = len(pedidos)
    fat_itens = sum_col(pedidos, "VALOR DOS ITENS (R$)", mask_concluido)
    pago_cliente_ped = sum_col(pedidos, "TOTAL PAGO PELO CLIENTE (R$)", mask_concluido)
    taxa_entrega_cliente = sum_col(pedidos, "TAXA DE ENTREGA PAGA PELO CLIENTE (R$)", mask_concluido)
    incentivo_ifood_ped = sum_col(pedidos, "INCENTIVO PROMOCIONAL DO IFOOD (R$)", mask_concluido)
    incentivo_loja_ped = sum_col(pedidos, "INCENTIVO PROMOCIONAL DA LOJA (R$)", mask_concluido)
    taxa_servico_ped = sum_col(pedidos, "TAXA DE SERVIÇO (R$)", mask_concluido)
    taxas_comissoes_ped = abs(sum_col(pedidos, "TAXAS E COMISSOES (R$)", mask_concluido))
    valor_liquido_ped = sum_col(pedidos, "VALOR LIQUIDO (R$)", mask_concluido)

    # Financeiro: usa as descrições do extrato de conciliação
    entrada_financeira = sum_financeiro(financeiro, ["Entrada Financeira"], absolute=False)
    incentivos_ifood = sum_financeiro(financeiro, ["Promoção custeada pelo iFood"], absolute=True)
    ressarcimentos = sum_financeiro(financeiro, ["Ressarcimento"], absolute=True)
    promocoes_loja = sum_financeiro(financeiro, ["Promoção custeada pela loja"], absolute=True)
    anuncios = sum_financeiro(financeiro, ["pacote de anúncios"], absolute=True)
    comissoes = sum_financeiro(financeiro, ["Comissão"], absolute=True)
    taxas = sum_financeiro(financeiro, ["Taxa entrega iFood", "Taxa de transação", "Taxa de serviço iFood", "Taxa de conveniência"], absolute=True)

    # Repasse líquido: soma geral do financeiro, mas exclui linhas que não impactam o repasse se a coluna existir
    if "impacto_no_repasse" in financeiro.columns and "valor" in financeiro.columns:
        impacto = financeiro["impacto_no_repasse"].fillna("").astype(str).str.upper().str.strip()
        base_repasse = financeiro.loc[impacto.eq("SIM"), "valor"]
        repasse_liquido = float(to_number(base_repasse).sum())
    elif "valor" in financeiro.columns:
        repasse_liquido = float(to_number(financeiro["valor"]).sum())
    else:
        repasse_liquido = 0.0

    dados_desempenho = montar_desempenho(desempenho)

    ticket_medio = fat_itens / pedidos_concluidos if pedidos_concluidos else 0
    dias_periodo = 30
    pedidos_dia = pedidos_concluidos / dias_periodo
    fat_dia = fat_itens / dias_periodo
    fat_comercial_dia = dados_desempenho["Faturamento comercial iFood"] / dias_periodo if dados_desempenho["Faturamento comercial iFood"] else 0
    repasse_dia = repasse_liquido / dias_periodo
    repasse_pedido = repasse_liquido / pedidos_concluidos if pedidos_concluidos else 0
    diferenca_comercial_itens = dados_desempenho["Faturamento comercial iFood"] - fat_itens if dados_desempenho["Faturamento comercial iFood"] else 0

    resumo = {
        "Pedidos totais": pedidos_totais,
        "Pedidos concluídos": pedidos_concluidos,
        "Pedidos cancelados": pedidos_cancelados,
        "Cancelamentos parciais": cancelamentos_parciais,
        "Ticket médio operacional": ticket_medio,
        "Faturamento bruto operacional (itens)": fat_itens,
        "Faturamento comercial iFood": dados_desempenho["Faturamento comercial iFood"],
        "Pedidos desempenho": dados_desempenho["Pedidos desempenho"],
        "Ticket médio desempenho": dados_desempenho["Ticket médio desempenho"],
        "Taxa de entrega desempenho": dados_desempenho["Taxa de entrega desempenho"],
        "Novos clientes": dados_desempenho["Novos clientes"],
        "Diferença comercial x itens": diferenca_comercial_itens,
        "Valor pago pelos clientes (pedidos)": pago_cliente_ped,
        "Venda recebida dos clientes (financeiro)": entrada_financeira,
        "Taxa de entrega paga pelos clientes": taxa_entrega_cliente,
        "Incentivos iFood (financeiro)": incentivos_ifood,
        "Incentivos iFood (pedidos)": incentivo_ifood_ped,
        "Promoções da loja (financeiro)": promocoes_loja,
        "Promoções da loja (pedidos)": incentivo_loja_ped,
        "Ressarcimentos": ressarcimentos,
        "Anúncios iFood": anuncios,
        "Comissões iFood": comissoes,
        "Taxas iFood": taxas,
        "Taxa de serviço (pedidos)": taxa_servico_ped,
        "Taxas e comissões (pedidos)": taxas_comissoes_ped,
        "Valor líquido (pedidos)": valor_liquido_ped,
        "Repasse líquido recebido": repasse_liquido,
        "Pedidos/dia": pedidos_dia,
        "Faturamento/dia": fat_dia,
        "Faturamento comercial/dia": fat_comercial_dia,
        "Repasse/dia": repasse_dia,
        "Repasse/pedido": repasse_pedido,
        "Promoções % fat. operacional": (promocoes_loja / fat_itens * 100) if fat_itens else 0,
        "Anúncios % fat. operacional": (anuncios / fat_itens * 100) if fat_itens else 0,
        "Comissões % fat. operacional": (comissoes / fat_itens * 100) if fat_itens else 0,
        "Taxas % fat. operacional": (taxas / fat_itens * 100) if fat_itens else 0,
        "Promoções + anúncios % fat. operacional": ((promocoes_loja + anuncios) / fat_itens * 100) if fat_itens else 0,
        "Repasse líquido % fat. operacional": (repasse_liquido / fat_itens * 100) if fat_itens else 0,
    }
    return resumo

# =========================
# Interface
# =========================
st.title("🍔 Resumo iFood - GGK / Sampa Burger")
st.caption("Suba os relatórios do iFood para gerar o resumo consolidado com faturamento comercial, operacional e financeiro.")

with st.expander("📖 Como usar", expanded=True):
    st.markdown("""
### 1. Relatório de Pedidos iFood
Entre no Portal do iFood:

**Menu → Pedidos → Selecione o período desejado → Exportar**

### 2. Relatório Financeiro / Conciliação iFood
Entre no Portal do iFood:

**Financeiro → Selecione o mês desejado → Exportar**

### 3. Relatório de Desempenho / Vendas iFood
Entre no Portal do iFood:

**Desempenho → Vendas → Selecione o período desejado → Exportar**

### 4. Atenção aos períodos
Os períodos analisados precisam ser exatamente os mesmos para que a comparação fique correta.

Exemplo correto: **Pedidos Abril/2026 + Financeiro Abril/2026 + Desempenho Abril/2026** ✅

Exemplo incorreto: **Pedidos Abril/2026 + Financeiro Maio/2026 + Desempenho Abril/2026** ❌
""")

col1, col2, col3 = st.columns(3)
with col1:
    arq_pedidos = st.file_uploader("Relatório de Pedidos iFood (.xlsx)", type=["xlsx"], key="pedidos")
with col2:
    arq_financeiro = st.file_uploader("Relatório Financeiro / Conciliação iFood (.xlsx)", type=["xlsx"], key="financeiro")
with col3:
    arq_desempenho = st.file_uploader("Relatório de Desempenho / Vendas iFood (.xlsx)", type=["xlsx"], key="desempenho")

if arq_pedidos and arq_financeiro and arq_desempenho:
    try:
        pedidos = read_ifood_xlsx(arq_pedidos)
        financeiro = read_ifood_xlsx(arq_financeiro)
        desempenho = read_ifood_xlsx(arq_desempenho, sheet_name="Vendas")
        resumo = montar_resumo(pedidos, financeiro, desempenho)

        st.success("Resumo gerado com sucesso.")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Faturamento comercial iFood", money(resumo["Faturamento comercial iFood"]))
        c2.metric("Faturamento operacional (itens)", money(resumo["Faturamento bruto operacional (itens)"]))
        c3.metric("Pedidos concluídos", num(resumo["Pedidos concluídos"]))
        c4.metric("Repasse líquido", money(resumo["Repasse líquido recebido"]))

        st.subheader("📊 Painel Consolidado")
        painel = pd.DataFrame([
            ["Faturamento comercial iFood (aba Desempenho)", money(resumo["Faturamento comercial iFood"])],
            ["Faturamento bruto operacional (itens vendidos)", money(resumo["Faturamento bruto operacional (itens)"])],
            ["Diferença: Comercial iFood - Itens", money(resumo["Diferença comercial x itens"])],
            ["Pedidos totais", resumo["Pedidos totais"]],
            ["Pedidos concluídos", resumo["Pedidos concluídos"]],
            ["Pedidos desempenho", resumo["Pedidos desempenho"]],
            ["Pedidos cancelados", resumo["Pedidos cancelados"]],
            ["Cancelamentos parciais", resumo["Cancelamentos parciais"]],
            ["Ticket médio operacional (itens / pedidos concluídos)", money(resumo["Ticket médio operacional"])],
            ["Ticket médio comercial iFood", money(resumo["Ticket médio desempenho"])],
            ["Novos clientes", resumo["Novos clientes"]],
            ["Valor pago pelos clientes (relatório de pedidos)", money(resumo["Valor pago pelos clientes (pedidos)"])],
            ["Venda recebida dos clientes (financeiro)", money(resumo["Venda recebida dos clientes (financeiro)"])],
            ["Taxa de entrega paga pelos clientes (pedidos)", money(resumo["Taxa de entrega paga pelos clientes"])],
            ["Taxa de entrega considerada no desempenho", money(resumo["Taxa de entrega desempenho"])],
            ["Incentivos iFood", money(resumo["Incentivos iFood (financeiro)"])],
            ["Ressarcimentos", money(resumo["Ressarcimentos"])],
            ["Promoções subsidiadas pela loja", money(resumo["Promoções da loja (financeiro)"])],
            ["Pacote de anúncios iFood", money(resumo["Anúncios iFood"])],
            ["Comissões iFood", money(resumo["Comissões iFood"])],
            ["Taxas iFood", money(resumo["Taxas iFood"])],
            ["Repasse líquido recebido", money(resumo["Repasse líquido recebido"])],
        ], columns=["Indicador", "Valor"])
        st.dataframe(painel, use_container_width=True, hide_index=True)

        st.info("Uso recomendado: faturamento comercial iFood para meta/ranking/royalties se esta for a regra da rede; faturamento operacional de itens para CMV, compras e produtividade; repasse líquido para caixa.")

        st.subheader("📈 Indicadores Gerenciais")
        kpis = pd.DataFrame([
            ["Pedidos por dia", num(resumo["Pedidos/dia"])],
            ["Faturamento comercial por dia", money(resumo["Faturamento comercial/dia"])],
            ["Faturamento operacional por dia", money(resumo["Faturamento/dia"])],
            ["Repasse líquido por dia", money(resumo["Repasse/dia"])],
            ["Repasse líquido por pedido", money(resumo["Repasse/pedido"])],
            ["Promoções sobre faturamento operacional", pct(resumo["Promoções % fat. operacional"])],
            ["Anúncios sobre faturamento operacional", pct(resumo["Anúncios % fat. operacional"])],
            ["Comissões sobre faturamento operacional", pct(resumo["Comissões % fat. operacional"])],
            ["Taxas sobre faturamento operacional", pct(resumo["Taxas % fat. operacional"])],
            ["Investimento em crescimento (promoções + anúncios)", pct(resumo["Promoções + anúncios % fat. operacional"])],
            ["Repasse líquido sobre faturamento operacional", pct(resumo["Repasse líquido % fat. operacional"])],
        ], columns=["KPI", "Valor"])
        st.dataframe(kpis, use_container_width=True, hide_index=True)

        st.subheader("💰 Visão Financeira Simplificada")
        receita_apos_promocoes = resumo["Faturamento bruto operacional (itens)"] - resumo["Promoções da loja (financeiro)"]
        visao = pd.DataFrame([
            ["Faturamento comercial iFood (Desempenho)", money(resumo["Faturamento comercial iFood"])],
            ["Faturamento bruto operacional (itens)", money(resumo["Faturamento bruto operacional (itens)"])],
            ["(-) Promoções da loja", money(resumo["Promoções da loja (financeiro)"])],
            ["Receita após promoções", money(receita_apos_promocoes)],
            ["(+) Incentivos iFood", money(resumo["Incentivos iFood (financeiro)"])],
            ["(+) Ressarcimentos", money(resumo["Ressarcimentos"])],
            ["(-) Comissões", money(resumo["Comissões iFood"])],
            ["(-) Taxas", money(resumo["Taxas iFood"])],
            ["(-) Anúncios", money(resumo["Anúncios iFood"])],
            ["Repasse líquido final", money(resumo["Repasse líquido recebido"])],
        ], columns=["Etapa", "Valor"])
        st.dataframe(visao, use_container_width=True, hide_index=True)

        whatsapp = f"""📊 Resumo iFood

📈 Faturamento Comercial iFood: {money(resumo['Faturamento comercial iFood'])}
🍔 Faturamento Operacional (Itens): {money(resumo['Faturamento bruto operacional (itens)'])}
📦 Pedidos concluídos: {num(resumo['Pedidos concluídos'])}
🎟️ Ticket médio operacional: {money(resumo['Ticket médio operacional'])}
🎟️ Ticket médio comercial iFood: {money(resumo['Ticket médio desempenho'])}

💸 Promoções da loja: {money(resumo['Promoções da loja (financeiro)'])}
📢 Anúncios: {money(resumo['Anúncios iFood'])}
🏦 Comissões: {money(resumo['Comissões iFood'])}
📋 Taxas: {money(resumo['Taxas iFood'])}

🤝 Incentivos iFood: {money(resumo['Incentivos iFood (financeiro)'])}
🔄 Ressarcimentos: {money(resumo['Ressarcimentos'])}

✅ Repasse líquido recebido: {money(resumo['Repasse líquido recebido'])}

📌 Diferença Comercial iFood x Itens: {money(resumo['Diferença comercial x itens'])}
📈 Investimento em crescimento (promoções + anúncios): {money(resumo['Promoções da loja (financeiro)'] + resumo['Anúncios iFood'])} ({pct(resumo['Promoções + anúncios % fat. operacional'])} do faturamento operacional)"""
        st.subheader("📱 Resumo para WhatsApp")
        st.text_area("Copiar e colar", whatsapp, height=310)

        csv = pd.concat([painel, kpis.rename(columns={"KPI":"Indicador"}), visao.rename(columns={"Etapa":"Indicador"})], ignore_index=True)
        st.download_button("Baixar resumo em CSV", csv.to_csv(index=False).encode("utf-8-sig"), file_name="resumo_ifood.csv", mime="text/csv")

        with st.expander("Conferência técnica"):
            st.write(f"Linhas lidas no relatório de pedidos: {len(pedidos)}")
            st.write(f"Linhas lidas no financeiro: {len(financeiro)}")
            st.write(f"Linhas lidas no desempenho/vendas: {len(desempenho)}")
            st.write("Observação: para CMV, compras e produtividade, use o faturamento bruto operacional de itens. Para comparar com a aba Desempenho do iFood, use o faturamento comercial iFood.")

    except Exception as e:
        st.error("Não consegui processar os arquivos. Confira se os três arquivos são relatórios XLSX do iFood.")
        st.exception(e)
else:
    st.info("Aguardando upload dos 3 arquivos.")
